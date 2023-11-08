import sys
from tkinter.filedialog import askopenfile

from openpyxl.reader.excel import load_workbook

global open_filename
global open_filename2

file = askopenfile(mode='r', filetypes=[
        ('Excel Files', '*.xlsx *.xlsm *.sxc *.ods *.csv *.tsv')])  # To open the file that you want.


wb = load_workbook(filename=file.name)  # Load into openpyxl
wb2 = wb.active

sheets = wb.sheetnames
sh1 = wb[sheets[0]]
sh2 = wb[sheets[1]]
sh3 = wb[sheets[2]]
open_filename = sh1
open_filename2 = sh2
open_filename3 = sh3



row = sh1.max_row
column = sh1.max_column

row2 = sh2.max_row
column2 = sh2.max_column

row3 = sh3.max_row
column3 = sh3.max_column

batch_seq = [[] for i in range(column)]

for i in range(1, column + 1):

    for j in range(2, row + 1):
        # print(sh1.cell(i,1).value)
        if sh1.cell(j, i).value != None:
            batch_seq[i - 1].append(sh1.cell(j, i).value)


print(row2)
print(column2)


for i in batch_seq:
    print(i)


volume = [0 for i in range(row2-1)]
name = [None for i in range(row2-1)]
active = [False for i in range(row2-1)]

for i in range(1, row2):
    volume[i-1] = sh2.cell(i+1, 3).value ##third column in sheet
    name[i-1] = sh2.cell(i+1, 1).value  ##first column in sheet
    active[i - 1] = sh2.cell(i+1, 2).value  ##second column in sheet


process_times = [[] for i in range(column3)]

for i in range(1, column3 + 1):

    for j in range(2, row3 + 1):
        # print(sh1.cell(i,1).value)
        if sh3.cell(j, i).value != None:
            process_times[i - 1].append(sh3.cell(j, i).value)


print(volume)
print(name)
print(active)

print(process_times)


