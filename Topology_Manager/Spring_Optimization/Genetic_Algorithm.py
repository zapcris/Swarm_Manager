from dataclasses import dataclass
from datetime import datetime
import numpy as np
import networkx as nx
import matplotlib.pyplot as plt
import random

import pymongo
from sklearn import preprocessing
from sklearn.preprocessing import MinMaxScaler
# from .utils import Get_Distance_Or_Flow
from scipy.sparse import csr_matrix
from scipy.sparse.csgraph import connected_components
from scipy.sparse import csr_array
from itertools import groupby
from collections import Counter
from networkx.algorithms import approximation
import sys
import xlrd
from tkinter import *
from tkinter.filedialog import askopenfile
from openpyxl import load_workbook

from UI import open_file
from batch_topology import create_batch_topology, unique_values_in_list_of_lists
from production_performance import prod_efficiency

from variant_topology import config, topology, workstation
import networkx as nx
from tkinter import *
from tkinter.filedialog import askopenfile
from openpyxl import load_workbook
from ttkbootstrap.constants import *
import pandas as pd

random.seed(1314141)


@dataclass
class chromosome:
    k_val: float
    iter_nr: int
    sequence: list


def max_value(input_list):
    return max([sublist[-1] for sublist in input_list])


def run_GA():
    global open_filename
    global open_filename2
    global open_filename3
    global open_filename4

    file = askopenfile(mode='r', filetypes=[
        ('Excel Files', '*.xlsx *.xlsm *.sxc *.ods *.csv *.tsv')])  # To open the file that you want.
    # ' mode='r' ' is to tell the filedialog to read the file
    # 'filetypes=[()]' is to filter the files shown as only Excel files

    wb = load_workbook(filename=file.name)  # Load into openpyxl
    wb2 = wb.active

    sheets = wb.sheetnames
    sh1 = wb[sheets[0]]
    sh2 = wb[sheets[1]]
    sh3 = wb[sheets[2]]
    sh4 = wb[sheets[3]]
    open_filename = sh1
    open_filename2 = sh2
    open_filename3 = sh3
    open_filename4 = sh4

    # print(wb.sheetnames)
    # print(sheets)

    row = sh1.max_row
    column = sh1.max_column

    row2 = sh2.max_row
    column2 = sh2.max_column

    row3 = sh3.max_row
    column3 = sh3.max_column

    row4 = sh4.max_row
    column4 = sh4.max_column

    batch_seq = [[] for i in range(column)]

    for i in range(1, column + 1):

        for j in range(2, row + 1):
            # print(sh1.cell(i,1).value)
            if sh1.cell(j, i).value != None:
                batch_seq[i - 1].append(sh1.cell(j, i).value)

    # batch_seq = [[1, 5, 9, 10, 2, 11, 13, 15, 7, 20],
    #              [1, 2, 7, 3, 5, 6, 8, 9, 13, 15, 19, 20],
    #              [1, 5, 8, 6, 3, 2, 4, 10, 15, 17, 20],
    #              [1, 8, 9, 10, 2, 11, 13, 15, 7, 20],
    #              [1, 4, 17, 3, 8, 9, 13, 15, 19, 20],
    #              [1, 6, 8, 6, 3, 12, 4, 10, 15, 17, 20],
    #              [1, 14, 8, 6, 13, 2, 4, 10, 15, 17, 20]]

    for i in batch_seq:
        print(i)

    prod_volume = [0 for i in range(row2 - 1)]
    prod_name = [None for i in range(row2 - 1)]
    prod_active = [False for i in range(row2 - 1)]
    process_times = [[] for i in range(column3)]
    wk_type = [0 for i in range(row4 - 1)]
    wk_capabilities = [[] for i in range(row4 - 1)]

    for i in range(1, row2):
        prod_volume[i - 1] = sh2.cell(i + 1, 3).value  ##third column in sheet
        prod_name[i - 1] = sh2.cell(i + 1, 1).value  ##first column in sheet
        prod_active[i - 1] = sh2.cell(i + 1, 2).value  ##second column in sheet

    for i in range(1, column3 + 1):
        for j in range(2, row3 + 1):
            # print(sh1.cell(i,1).value)
            if sh3.cell(j, i).value != None:
                process_times[i - 1].append(sh3.cell(j, i).value)

    for i in range(1, row4):
        wk_type[i - 1] = sh4.cell(i + 1, 1).value
        string_numbers = sh4.cell(i + 1, 2).value
        print(isinstance(string_numbers, int))
        if isinstance(string_numbers, int):
            wk_capabilities[i - 1] = [string_numbers, 99]
        else:
            wk_capabilities[i - 1] = list(map(int, string_numbers.split(',')))

    print("wk_capability", wk_capabilities)


    init_population = []
    start_k = 1.2
    stop_k = 2.0
    step_k = 0.2

    ## Calculate 2 sets of population with different iteration value 25 and 35#####

    for i in range(int(start_k * 10), int(stop_k * 10), int(step_k * 10)):
        chrm_1 = chromosome(i / 10, 10, batch_seq)
        init_population.append(chrm_1)
        # print(i / 10)

    for i in range(int(start_k * 10), int(stop_k * 10), int(step_k * 10)):
        chrm_2 = chromosome(i / 10, 15, batch_seq)
        init_population.append(chrm_2)

    # for p in population:
    #     print(p)

    fitness_list = []
    topology_htable = dict()
    for i in range(len(init_population)):
        btop = create_batch_topology(init_population[i].sequence, i + 1, init_population[i].k_val,
                                     init_population[i].iter_nr)
        fitness_list.append(btop[0])
        topology_htable.update({btop[0]: (btop[1], btop[2], init_population[i].k_val, init_population[i].iter_nr)})
        # print(btop)

    print("Fitness list:", fitness_list)

    # for key, value in topology_htable.items():
    #     print(key, value)

    ########choosing the parents ####################
    middle_index = round(len(init_population) / 2)

    sorted_fitness1 = fitness_list[:middle_index]
    sorted_fitness2 = fitness_list[middle_index:]

    print(sorted_fitness1.index(sorted(sorted_fitness1)[0]))
    print(sorted_fitness2.index(sorted(sorted_fitness2)[0]))

    parent_1 = init_population[sorted_fitness1.index(sorted(sorted_fitness1)[0])]
    parent_2 = init_population[sorted_fitness2.index(sorted(sorted_fitness2)[0]) + middle_index]

    print("Parent 1:", parent_1)
    print("Parent 2:,", parent_2)

    ####3 Crossover part of Genetic algorithm##################
    off_population = []
    offspring_fitness = []

    offspring_1 = chromosome(parent_1.k_val, parent_2.iter_nr, parent_1.sequence)
    off_population.append(offspring_1)

    offspring_2 = chromosome(parent_2.k_val, parent_1.iter_nr, parent_2.sequence)
    off_population.append(offspring_2)

    for i in range(len(off_population)):
        off_top = create_batch_topology(off_population[i].sequence, i + 1, off_population[i].k_val,
                                        off_population[i].iter_nr)
        offspring_fitness.append(off_top[0])
        topology_htable.update(
            {off_top[0]: (off_top[1], off_top[2], off_population[i].k_val, off_population[i].iter_nr)})
        # print("OFF spring topologies:", i + 1, off_top)

    print(min(offspring_fitness))

    # for key, value in topology_htable.items():
    #     print(key, value)

    ### Mutation function to be decided later#####3#
    mut_population = []
    mut_fitness = []

    ###Recursive operation until desired fitness achieved#############
    min_fitness = []

    def GA_recursion(itr1, itr2, rec_nr):
        new_population = []
        fit_list = []
        off_populn = []
        offspr_fitness = []
        start_k = 1.3
        stop_k = 2.0
        step_k = (stop_k - start_k) / 0.25
        print(f'The recursion number is {rec_nr} with iteration pari {itr1} and {itr2}')
        # print("Cleared length of population:", len(new_population))
        for i in range(int(start_k * 10), int(stop_k * 10), int(step_k)):
            chrm1 = chromosome(i / 10, itr1, batch_seq)
            new_population.append(chrm1)
            # print(i / 10)

        for i in range(int(start_k * 10), int(stop_k * 10), int(step_k)):
            chrm2 = chromosome(i / 10, itr2, batch_seq)
            new_population.append(chrm2)

        for i in range(len(new_population)):
            print("The new population:", new_population[i])
            top = create_batch_topology(new_population[i].sequence, i + 1, new_population[i].k_val,
                                        new_population[i].iter_nr)
            fit_list.append(top[0])
            topology_htable.update({top[0]: (top[1], top[2], new_population[i].k_val, new_population[i].iter_nr)})

        print("The current population fitness list:", fit_list)

        m_index = round(len(new_population) / 2)

        sorted_fit1 = fit_list[:m_index]
        sorted_fit2 = fit_list[m_index:]

        p_1 = new_population[sorted_fit1.index(sorted(sorted_fit1)[0])]
        print("The index of 1st parent in fit list:", sorted_fit1.index(sorted(sorted_fit1)[0]))

        p_2 = new_population[sorted_fit2.index(sorted(sorted_fit2)[0]) + m_index]
        print("The index of 2nd parent in fit list:", sorted_fit2.index(sorted(sorted_fit2)[0]) + m_index)

        offspr_1 = chromosome(p_1.k_val, p_2.iter_nr, p_1.sequence)
        off_populn.append(offspr_1)
        offspr_2 = chromosome(p_2.k_val, p_1.iter_nr, p_2.sequence)
        off_populn.append(offspr_2)

        print("offspring 1 chromosome:", offspr_1)
        print("offspring 2 chromosome:", offspr_2)

        for i in range(len(off_populn)):
            otop = create_batch_topology(off_population[i].sequence, i + 1, off_population[i].k_val,
                                         off_population[i].iter_nr)
            offspr_fitness.append(otop[0])
            topology_htable.update({otop[0]: (otop[1], otop[2], off_population[i].k_val, off_population[i].iter_nr)})
            # print("OFF spring topologies:", i + 1, otop)

        print("fitness list of offspring in this iteration:", offspr_fitness)
        print("minimum fitnesss value in this iteration:", min(offspr_fitness))

        if min(fit_list) < min(offspr_fitness):
            gen_min_fitness = min(fit_list)
        else:
            gen_min_fitness = min(offspr_fitness)
        print("minimumfitness value in this generation:", gen_min_fitness)

        # print(int(step_k))
        # print(result)

        if min(offspr_fitness) > 100 and itr1 != 40 and itr2 != 45:
            min_fitness.append(gen_min_fitness)
            # new_population.clear()
            # fit_list.clear()
            # off_populn.clear()
            # offspr_fitness.clear()
            # del p_1
            # del p_2
            # del offspr_1
            # del offspr_2

            GA_recursion(itr1 + 5, itr2 + 5, rec_nr + 1)
            result = min(min_fitness)

        elif min(offspr_fitness) <= 100:
            result = min(offspr_fitness)

        else:
            result = 0

        return result

    ### Draw the optimal topology#####
    OptmialGraph = nx.MultiGraph()
    n_list = unique_values_in_list_of_lists(batch_seq)
    e_list = []
    for i in range(len(batch_seq)):
        for j in range(len(batch_seq[i]) - 1):
            # print(graph[i][j], graph[i][j+1])
            edges = [batch_seq[i][j], batch_seq[i][j + 1]]
            e_list.append(edges)

    OptmialGraph.add_nodes_from(n_list)
    OptmialGraph.add_edges_from(e_list)

    ##### END of GA ######
    if min(offspring_fitness) > 0:
        print("\n\nRecursion Started")

        final_fitness = (GA_recursion(20, 25, 1))

        print("the min fitness list:", min_fitness)
        print("The least possible fitness value:", final_fitness)
        print("The topology of the fittest value:", topology_htable[final_fitness])

        nx.draw(OptmialGraph, topology_htable[final_fitness][0], with_labels=True)
        # plt.savefig('optimal topology found from GA recursion')
        plt.savefig("Optimized_Spring.pdf", format="pdf", bbox_inches="tight")
        plt.clf()

    elif min(offspring_fitness) <= 10:
        print("Fitness value found below 500:", min(offspring_fitness))
        nx.draw(OptmialGraph, topology_htable[min(offspring_fitness)][0], with_labels=True)
        plt.savefig("Optimized_Spring.pdf", format="pdf", bbox_inches="tight")
        plt.clf()

    width_dict = Counter(OptmialGraph.edges())
    edge_width = [[u, v, {'frequency': value}]
                  for ((u, v), value) in width_dict.items()]
    print("The frequency of edges", edge_width)

    "Production performance of the fittest solution"
    # Qty_order = [10, 30, 50, 20, 60, 20, 40, 30, 40, 20]
    Qty_order = prod_volume
    # Qty_order = [1, 1, 1, 1, 1, 1, 1]
    print("Positions of Optimal Topology", topology_htable[final_fitness][0])
    print(prod_efficiency(batch_seq, topology_htable[final_fitness][0], Qty_order, topology_htable[final_fitness][1]))

    top_keys = []
    for fit in topology_htable:
        k = fit
        top_keys.append(k)

    topologies = []
    print("New code")

    # print(topology_htable[final_fitness][0])
    # print(top_keys)

    for fit_val in top_keys:
        # topologies.append(topology_htable[fit_val][1])
        top_dict = topology_htable[fit_val][0]
        top = [None] * max_value(batch_seq)
        for key, value in top_dict.items():
            top[key - 1] = value
        topologies.append(top)

    optimized_top = [None] * max_value(batch_seq)

    ##old logic##
    # for key, value in topology_htable[final_fitness][0].items():
    #     optimized_top[key-1] = value
    # #print(topologies)

    for i in range(len(optimized_top)):
        for key, value in topology_htable[final_fitness][0].items():
            # print(key, value)
            if i == key - 1:
                # print(i)
                optimized_top[i] = list(dict.fromkeys(value))

    _time = datetime.now()
    current_dateTime = _time.strftime("%Y-%m-%d-%H-%M-%S")
    "Connect to MongoDB"
    client = pymongo.MongoClient("mongodb://localhost:27017")
    db = client["Topology_Manager"]
    collection = db["Spring_Topologies"]

    coll_dict = {"Timestamp": current_dateTime,
                 "Product_name": prod_name,
                 "Product_volume": prod_volume,
                 "Product_active": prod_active,
                 "Process_Sequence": batch_seq,
                 "Process_times": process_times,
                 "WK_type": wk_type,
                 "WK_capabilities" : wk_capabilities,
                 "Statistical_Fitness": top_keys,
                 "Estimated_Topologies": topologies, "Optimized_Topology": optimized_top}
    # coll_dict = {"Topologies": topologies}

    x = collection.insert_one(coll_dict)

    def save(Topology):
        "Connect to MongoDB"
        client = pymongo.MongoClient("mongodb://localhost:27017")
        db = client["Topology_Manager"]
        collection = db["Optimal_Topology"]

        coll_dict = {"Name": "Optimal_Spring", "Topology": Topology,
                     "Product_name": prod_name,
                     "Product_active": prod_active,
                     "Process_Sequence": batch_seq,
                     "Product_volume": prod_volume,
                     "Process_times": process_times,
                     "WK_type": wk_type,
                     "WK_capabilities": wk_capabilities}
        # coll_dict = {"Topologies": topologies}

        total_doc = collection.count_documents({})
        if total_doc == 0:
            collection.insert_one(coll_dict)
        else:
            collection.replace_one({"Name": "Optimal_Spring"}, coll_dict)

    ## Save Spring optimal topology for Tree Optimization####
    #save(optimized_top)
